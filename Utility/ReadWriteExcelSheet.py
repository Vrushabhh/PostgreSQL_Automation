'''
@author: Vrushabh_Hukkerikar
'''
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,colors,Font,NamedStyle



class ReadExcel(object):
    '''
    classdocs
    '''
    sheetname=""
    def GetCellData(self,ExcelFileLocation,Sheetname,column,row):
        wb = load_workbook(ExcelFileLocation,data_only=True)
        ReadExcel.sheetname = wb[Sheetname]
        CellLocation = column + row 
#         print("reading")

        #print(ReadExcel.sheetname[CellLocation].value)
        
        return (ReadExcel.sheetname[CellLocation].value)
    
    def GetRowCount(self,ExcelFileLocation,SN1):
        #print(SN1)
        Sheetname1 = SN1
        wb = load_workbook(ExcelFileLocation) 
        SN = wb[Sheetname1]
        row_count = SN.max_row
        return row_count 
    
    def WriteExcel(self,ExcelFileLocation,Sheetname,colno,rowno,storingvalue):
        wb = load_workbook(ExcelFileLocation,data_only=True)
        ReadExcel.sheetname = wb[Sheetname]
        #CellLocation = column + row 
        sheetname=wb.active
        NewCell=sheetname.cell(row=colno, column=rowno)
        NewCell.value=storingvalue
        wb.save(ExcelFileLocation)
        return (NewCell.value) 
    
    def GreenBgrColor(self,ExcelFileLocation,Sheetname,colno,rowno):
        wb = load_workbook(ExcelFileLocation,data_only=True)
        ReadExcel.sheetname = wb[Sheetname]
        sheetname=wb.active
        NewCell=sheetname.cell(row=colno, column=rowno)
        NewCell.fill = PatternFill(fgColor=colors.GREEN, fill_type="solid")
        wb.save(ExcelFileLocation)
        return (NewCell.fill)
    
    def RedBgrColor(self,ExcelFileLocation,Sheetname,colno,rowno):
        wb = load_workbook(ExcelFileLocation,data_only=True)
        ReadExcel.sheetname = wb[Sheetname]
        sheetname=wb.active
        NewCell=sheetname.cell(row=colno, column=rowno)
        NewCell.fill = PatternFill(fgColor=colors.RED, fill_type="solid")
        wb.save(ExcelFileLocation)
        return (NewCell.fill)
        
    
    def __init__(self):
        '''
        Constructor
        '''
